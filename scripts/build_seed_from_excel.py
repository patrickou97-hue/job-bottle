from __future__ import annotations

import csv
import re
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "data" / "source" / "27秋招信息整理.xlsx"
CSV_OUTPUT = ROOT / "data" / "processed" / "27_jobs_import.csv"
SQL_OUTPUT = ROOT / "supabase" / "seed.sql"


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def sql_string(value: str | None) -> str:
    if not value:
        return "null"
    return "'" + value.replace("'", "''") + "'"


def is_http_url(value: str) -> bool:
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def split_tags(*values: str | None) -> list[str]:
    tags: list[str] = []
    for value in values:
        for item in re.split(r"[,，、/|｜\s]+", value or ""):
            item = item.strip()
            if item and len(item) <= 16 and item not in tags:
                tags.append(item)
    return tags[:18]


def sql_array(values: list[str]) -> str:
    if not values:
        return "'{}'::text[]"
    return "array[" + ", ".join(sql_string(value) for value in values) + "]::text[]"


def main() -> None:
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    ws = wb["27秋招正式批+提前批"]

    rows = []
    skipped = []
    for row_number in range(4, ws.max_row + 1):
        company_name = clean(ws.cell(row_number, 1).value)
        start_date = clean(ws.cell(row_number, 2).value)
        industry = clean(ws.cell(row_number, 3).value)
        batch_type = clean(ws.cell(row_number, 4).value)
        job_titles = clean(ws.cell(row_number, 5).value)
        locations = clean(ws.cell(row_number, 6).value)
        apply_url = clean(ws.cell(row_number, 7).value)
        notes = "；".join(
            item
            for item in [clean(ws.cell(row_number, 8).value), clean(ws.cell(row_number, 9).value)]
            if item
        )

        if not company_name or not apply_url or not is_http_url(apply_url):
            skipped.append((row_number, company_name, apply_url))
            continue

        tags = split_tags(industry, batch_type, locations, job_titles)
        rows.append(
            {
                "company_name": company_name,
                "start_date": start_date,
                "industry": industry,
                "batch_type": batch_type,
                "job_titles": job_titles,
                "locations": locations,
                "apply_url": apply_url,
                "notes": notes,
                "tags": tags,
            }
        )

    CSV_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with CSV_OUTPUT.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=[
                "公司名称",
                "开启时间",
                "所在行业",
                "类型",
                "招聘岗位",
                "工作地点",
                "投递链接",
                "备注",
            ],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "公司名称": row["company_name"],
                    "开启时间": row["start_date"],
                    "所在行业": row["industry"],
                    "类型": row["batch_type"],
                    "招聘岗位": row["job_titles"],
                    "工作地点": row["locations"],
                    "投递链接": row["apply_url"],
                    "备注": row["notes"],
                }
            )

    values_sql = []
    for row in rows:
        values_sql.append(
            "("
            + ", ".join(
                [
                    sql_string(row["company_name"]),
                    sql_string(row["start_date"]),
                    sql_string(row["industry"]),
                    sql_string(row["batch_type"]),
                    sql_string(row["job_titles"]),
                    sql_string(row["locations"]),
                    sql_string(row["apply_url"]),
                    sql_string(row["notes"]),
                    "null",
                    sql_array(row["tags"]),
                    "true",
                ]
            )
            + ")"
        )

    SQL_OUTPUT.write_text(
        "-- Generated from data/source/27秋招信息整理.xlsx.\n"
        f"-- Valid rows: {len(rows)}. Skipped rows: {len(skipped)}.\n\n"
        "insert into public.jobs (\n"
        "  company_name,\n"
        "  start_date,\n"
        "  industry,\n"
        "  batch_type,\n"
        "  job_titles,\n"
        "  locations,\n"
        "  apply_url,\n"
        "  notes,\n"
        "  logo_url,\n"
        "  tags,\n"
        "  is_active\n"
        ")\n"
        "select *\n"
        "from (\n"
        "values\n"
        + ",\n".join(values_sql)
        + "\n"
        ") as incoming (\n"
        "  company_name,\n"
        "  start_date,\n"
        "  industry,\n"
        "  batch_type,\n"
        "  job_titles,\n"
        "  locations,\n"
        "  apply_url,\n"
        "  notes,\n"
        "  logo_url,\n"
        "  tags,\n"
        "  is_active\n"
        ")\n"
        "where not exists (\n"
        "  select 1\n"
        "  from public.jobs existing\n"
        "  where existing.company_name = incoming.company_name\n"
        "    and existing.apply_url = incoming.apply_url\n"
        ");\n",
        encoding="utf-8",
    )

    print(f"valid={len(rows)} skipped={len(skipped)}")


if __name__ == "__main__":
    main()
